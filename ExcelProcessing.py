import pandas as pd
from openpyxl import load_workbook
file_path = "./files/Active_Contact_List.xlsx"

print("================ Reading file and print =============================")

# Reading the Excel file
df = pd.read_excel(file_path, sheet_name="01302011")

# Displaying the first few rows
print(df.head())

print("================= Modifying Cell Value ============================")

# Load an existing Excel file
workbook = load_workbook(file_path)

sheet = workbook["01302011"]

# Display column names with coordinates
for cell in sheet[1]:  # Iterate over the first row
    print(f"Value: {cell.value}, Coordinate: {cell.coordinate}")

# Modify a specific cell
sheet["A1"] = "Status"
# Save the changes
workbook.save(file_path)


print("================= Adding new sheet ============================")

# Add a new sheet to the workbook
workbook.create_sheet(title="Reports")

# Save the changes
workbook.save(file_path)

print("Created a new sheet")

print("================= Report generation ============================")


# Step 1: Read data from Excel file

# Step 2: Perform data processing (e.g., grouping and aggregating sales by region)
report = df.groupby("Company Name")["Status"].count().reset_index()

with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    # Add the report to a new sheet called "reports"
    report.to_excel(writer, sheet_name="Reports", index=False)

print("Report generated and saved to the 'reports' sheet in the same Excel file.")